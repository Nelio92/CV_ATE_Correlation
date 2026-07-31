from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from cv_ate_correlation.chip_manifest import (
    CHIP_MANIFEST_COLUMNS,
    CHIP_MANIFEST_FILENAME,
    CHIP_MANIFEST_SHEET,
    chip_manifest_template_path,
    create_chip_manifest_template,
    save_chip_manifest_template,
)
from cv_ate_correlation.extraction import read_chip_manifest


def test_packaged_chip_manifest_is_a_guided_one_sheet_template() -> None:
    path = chip_manifest_template_path()

    assert path.name == CHIP_MANIFEST_FILENAME
    assert path.is_file()
    workbook = load_workbook(path, data_only=False)
    try:
        assert workbook.sheetnames == [CHIP_MANIFEST_SHEET]
        sheet = workbook[CHIP_MANIFEST_SHEET]
        assert tuple(sheet.cell(row=1, column=column).value for column in range(1, 6)) == (
            CHIP_MANIFEST_COLUMNS
        )
        assert sheet.freeze_panes == "A2"
        assert sheet.tables["CorreLaTEChipManifest"].ref == "A1:E2"
        assert len(sheet.data_validations.dataValidation) == 5
        assert all(sheet.cell(row=1, column=column).comment for column in range(1, 6))
        assert "All five columns are REQUIRED" in sheet["G4"].value
        assert "TE–CV alignment" in sheet["K7"].value
        assert "process-variation" in sheet["K11"].value
        assert len(sheet._images) == 1
    finally:
        workbook.close()


def test_template_generator_and_distribution_copy_preserve_structure(tmp_path: Path) -> None:
    generated = create_chip_manifest_template(tmp_path / "generated.xlsx")
    distributed = save_chip_manifest_template(tmp_path / "distributed.xlsx")

    assert distributed.read_bytes() == chip_manifest_template_path().read_bytes()
    workbook = load_workbook(generated, read_only=False)
    try:
        assert workbook.sheetnames == [CHIP_MANIFEST_SHEET]
        assert workbook[CHIP_MANIFEST_SHEET].tables["CorreLaTEChipManifest"].ref == "A1:E2"
    finally:
        workbook.close()


def test_user_populated_packaged_template_is_ready_for_section_2(tmp_path: Path) -> None:
    path = save_chip_manifest_template(tmp_path / CHIP_MANIFEST_FILENAME)
    workbook = load_workbook(path)
    sheet = workbook[CHIP_MANIFEST_SHEET]
    for column, value in enumerate((1773, 2, 31, 5, "TT"), start=1):
        sheet.cell(row=2, column=column, value=value)
    workbook.save(path)
    workbook.close()

    chips, metadata = read_chip_manifest(path)

    assert chips == {("2", 31, 5)}
    assert metadata == {("2", 31, 5): {"DUT Nr": "1773", "DoE split": "TT"}}


def test_manifest_reader_accepts_complete_aliases_and_normalizes_numeric_wafer(
    tmp_path: Path,
) -> None:
    path = tmp_path / "chips.xlsx"
    pd.DataFrame({
        "DUT Number": [1773, 1774],
        "WAF": [2.0, 9.0],
        "Die X": [31, 35],
        "Die Y": [5, 54],
        "Process Corner": ["TT", "SS"],
    }).to_excel(path, index=False)

    chips, metadata = read_chip_manifest(path)

    assert chips == {("2", 31, 5), ("9", 35, 54)}
    assert metadata[("2", 31, 5)] == {"DUT Nr": "1773", "DoE split": "TT"}
    assert metadata[("9", 35, 54)] == {"DUT Nr": "1774", "DoE split": "SS"}


@pytest.mark.parametrize("missing_column", CHIP_MANIFEST_COLUMNS)
def test_manifest_reader_requires_every_canonical_field(
    tmp_path: Path,
    missing_column: str,
) -> None:
    values = {
        "DUT Nr": [1773],
        "Wafer": [2],
        "X": [31],
        "Y": [5],
        "DoE split": ["TT"],
    }
    values.pop(missing_column)
    path = tmp_path / f"missing-{missing_column}.xlsx"
    pd.DataFrame(values).to_excel(path, index=False)

    with pytest.raises(ValueError, match=rf"missing required column\(s\):.*{missing_column}"):
        read_chip_manifest(path)


@pytest.mark.parametrize("blank_column", CHIP_MANIFEST_COLUMNS)
def test_manifest_reader_rejects_blank_required_values(
    tmp_path: Path,
    blank_column: str,
) -> None:
    values = {
        "DUT Nr": [1773],
        "Wafer": [2],
        "X": [31],
        "Y": [5],
        "DoE split": ["TT"],
    }
    values[blank_column] = [None]
    path = tmp_path / f"blank-{blank_column}.xlsx"
    pd.DataFrame(values).to_excel(path, index=False)

    with pytest.raises(ValueError, match=rf"missing {blank_column}"):
        read_chip_manifest(path)


@pytest.mark.parametrize(
    ("values", "message"),
    [
        (
            {
                "DUT Nr": [1773, 1773],
                "Wafer": [2, 9],
                "X": [31, 35],
                "Y": [5, 54],
                "DoE split": ["TT", "SS"],
            },
            "duplicate DUT Nr",
        ),
        (
            {
                "DUT Nr": [1773, 1774],
                "Wafer": [2, 2],
                "X": [31, 31],
                "Y": [5, 5],
                "DoE split": ["TT", "TT"],
            },
            "duplicate Wafer/X/Y",
        ),
    ],
)
def test_manifest_reader_rejects_ambiguous_chip_identities(
    tmp_path: Path,
    values: dict[str, list[object]],
    message: str,
) -> None:
    path = tmp_path / "duplicates.xlsx"
    pd.DataFrame(values).to_excel(path, index=False)

    with pytest.raises(ValueError, match=message):
        read_chip_manifest(path)


def test_blank_packaged_template_gives_actionable_section_2_error() -> None:
    with pytest.raises(ValueError, match="contains no chip rows.*DUT Nr.*DoE split"):
        read_chip_manifest(chip_manifest_template_path())