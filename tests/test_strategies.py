from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from cv_ate_correlation.models import (
    CorrelationProfile,
    CovariateProfile,
    GuardBandProfile,
    RequirementRule,
    TestPolicy as CorrelationTestPolicy,
    TestSelector as ProfileTestSelector,
)
from cv_ate_correlation.correlation import (
    attach_covariate,
    attach_covariate_from_test_rows,
    correlate_frame,
)
from cv_ate_correlation.excel import ACCENT_1_BLUE, DATA_INTEREST_FILL
from cv_ate_correlation.reporting import write_excel_report, write_plots


def test_mean_delta_and_worst_case_upper_limit() -> None:
    frame = pd.DataFrame({
        "group": ["A"] * 5,
        "reference": [2, 3, 4, 5, 6],
        "ate": [1, 1, 3, 3, 5],
        "high": [10] * 5,
    })
    profile = CorrelationProfile(
        name="generic", strategy="Mean_Deltas", reference_column="reference", candidate_column="ate",
        group_by=("group",), minimum_points=5, upper_limit_column="high",
        guard_band=GuardBandProfile(kind="shifted_upper_limit"),
    )
    row = correlate_frame(frame, profile).summary.iloc[0]
    assert row["CorrelationFactor"] == pytest.approx(1.4)
    assert row["MaxAbsResidual"] == pytest.approx(0.6)
    assert row["AdjustedUpperLimit"] == pytest.approx(8.6)
    assert row["WorstCaseUpperLimit"] == pytest.approx(8.0)


def test_mean_and_median_delta_models_center_their_residuals_with_fixed_slope_one() -> None:
    frame = pd.DataFrame({
        "group": ["A"] * 6,
        "reference": [9.8, 10.2, 10.0, 9.7, 10.3, 10.1],
        "ate": [7.0, 7.5, 8.0, 8.5, 9.0, 9.5],
    })
    profile = CorrelationProfile(
        name="offset invariants",
        strategy="Mean_Deltas",
        reference_column="reference",
        candidate_column="ate",
        group_by=("group",),
        minimum_points=5,
    )

    result = correlate_frame(frame, profile)
    summary = result.summary.iloc[0]
    details = result.details
    direct_delta = frame["reference"] - frame["ate"]

    assert summary["MeanDelta"] == pytest.approx(direct_delta.mean())
    assert summary["MedianDelta"] == pytest.approx(direct_delta.median())
    assert details["MeanDeltasCorrectedCandidate"].tolist() == pytest.approx(
        (frame["ate"] + direct_delta.mean()).tolist()
    )
    assert details["MedianDeltasCorrectedCandidate"].tolist() == pytest.approx(
        (frame["ate"] + direct_delta.median()).tolist()
    )
    assert details["MeanDeltasResidual"].mean() == pytest.approx(0.0, abs=1e-12)
    assert details["MedianDeltasResidual"].median() == pytest.approx(0.0, abs=1e-12)


def test_median_offset_sigma_and_requirement_rule() -> None:
    frame = pd.DataFrame({
        "group": ["A"] * 5, "mode": ["special"] * 5,
        "reference": [10, 11, 12, 13, 20], "ate": [9, 10, 11, 12, 12],
    })
    profile = CorrelationProfile(
        name="generic", strategy="median_offset", reference_column="reference", candidate_column="ate",
        group_by=("group", "mode"), minimum_points=5,
        guard_band=GuardBandProfile(kind="distribution_sigma", rules=(RequirementRule(
            when={"mode": ("special",)}, lower=0, upper=20,
            lower_residual="maximum", upper_residual="minimum",
        ),)),
    )
    row = correlate_frame(frame, profile).summary.iloc[0]
    assert row["CorrelationFactor"] == pytest.approx(1.0)
    assert row["ResidualMax"] == pytest.approx(7.0)
    assert row["ResidualMin"] == pytest.approx(0.0)
    assert row["AdjustedLowerLimit"] == pytest.approx(7.0)
    assert row["AdjustedUpperLimit"] == pytest.approx(20.0)


def test_different_test_sets_use_different_strategies_and_guard_bands(tmp_path: Path) -> None:
    frame = pd.DataFrame({
        "group": ["A"] * 10,
        "Test Number": [101] * 5 + [202] * 5,
        "Test Name": ["Mean test"] * 5 + ["Median test"] * 5,
        "reference": [2, 3, 4, 5, 9, 2, 3, 4, 5, 9],
        "ate": [1, 2, 3, 4, 5, 1, 2, 3, 4, 5],
        "high": [10] * 10,
    })
    profile = CorrelationProfile(
        name="mixed policies",
        strategy="Mean_Deltas",
        reference_column="reference",
        candidate_column="ate",
        group_by=("group",),
        minimum_points=5,
        upper_limit_column="high",
        guard_band=GuardBandProfile(kind="shifted_upper_limit"),
        test_policies=(
            CorrelationTestPolicy(
                "Mean policy",
                ProfileTestSelector(exact=(101,)),
                "Mean_Deltas",
                GuardBandProfile(kind="shifted_upper_limit"),
            ),
            CorrelationTestPolicy(
                "Median policy",
                ProfileTestSelector(exact=(202,)),
                "median_offset",
                GuardBandProfile(kind="distribution_sigma", sigma_multiplier=4),
            ),
        ),
    )

    result = correlate_frame(frame, profile)
    summary = result.summary.set_index("TestSet")

    assert summary.loc["Mean policy", "CorrelationFactor"] == pytest.approx(1.6)
    assert summary.loc["Mean policy", "CorrelationStrategy"] == "Mean_Deltas"
    assert summary.loc["Mean policy", "GuardBandPolicy"] == "shifted_upper_limit"
    assert summary.loc["Median policy", "CorrelationFactor"] == pytest.approx(1.0)
    assert summary.loc["Median policy", "CorrelationStrategy"] == "Median_Deltas"
    assert summary.loc["Median policy", "GuardBandPolicy"] == "distribution_sigma"
    report = tmp_path / "mixed-policies.xlsx"
    write_excel_report(result, profile, report)
    factors = pd.read_excel(report, sheet_name="Correlation_Factors").set_index("TestSet")
    guards = pd.read_excel(report, sheet_name="Guard_Bands").set_index("TestSet")
    for output in (factors, guards):
        assert {"CorrelationStrategy", "GuardBandPolicy"}.issubset(output.columns)
    assert factors.loc["Mean policy", "CorrelationFactor"] == pytest.approx(1.6)
    assert factors.loc["Median policy", "CorrelationFactor"] == pytest.approx(1.0)
    assert not {
        "CorrelationFactorA", "CorrelationFactorB", "MeanDelta", "MedianDelta",
        "LinearSlope", "LinearIntercept", "PhysicsAlpha", "PhysicsBeta",
        "LinearR2", "MeanDeltasR2", "MedianDeltasR2", "PhysicsR2",
    }.intersection(factors.columns)
    assert pd.isna(guards.loc["Mean policy", "AdjustedLowerLimit"])
    assert guards.loc["Mean policy", "AdjustedUpperLimit"] == pytest.approx(8.4)
    assert guards.loc["Mean policy", "WorstCaseUpperLimit"] == pytest.approx(6.0)
    assert pd.notna(guards.loc["Median policy", "AdjustedLowerLimit"])
    assert pd.notna(guards.loc["Median policy", "AdjustedUpperLimit"])
    assert pd.isna(guards.loc["Median policy", "WorstCaseUpperLimit"])
    assert not {
        "RequirementMin", "RequirementMax", "OriginalLowerLimit", "OriginalUpperLimit",
    }.intersection(guards.columns)
    workbook = load_workbook(report)
    for worksheet in workbook.worksheets:
        assert worksheet.auto_filter.ref == worksheet.dimensions
        assert worksheet["A1"].fill.fgColor.rgb == f"00{ACCENT_1_BLUE}"
    for sheet_name, highlighted in {
        "Correlation_Factors": {"CorrelationFactor"},
        "Guard_Bands": {
            "GuardBandMethod", "AdjustedLowerLimit", "AdjustedUpperLimit", "WorstCaseUpperLimit",
        },
    }.items():
        worksheet = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in worksheet[1]}
        for name in highlighted:
            column = headers[name]
            populated = [
                worksheet.cell(row=row, column=column)
                for row in range(2, worksheet.max_row + 1)
                if worksheet.cell(row=row, column=column).value is not None
            ]
            assert populated
            assert all(cell.fill.fgColor.rgb == f"00{DATA_INTEREST_FILL}" for cell in populated)
            assert all(cell.font.bold for cell in populated)
        count_cell = worksheet.cell(row=2, column=headers["Count"])
        assert count_cell.fill.fill_type is None


def test_overlapping_test_set_policies_are_rejected_at_runtime() -> None:
    frame = pd.DataFrame({
        "group": ["A"] * 2,
        "Test Number": [101] * 2,
        "Test Name": ["Leakage"] * 2,
        "reference": [2, 3],
        "ate": [1, 2],
    })
    profile = CorrelationProfile(
        name="overlap",
        strategy="mean_delta",
        reference_column="reference",
        candidate_column="ate",
        group_by=("group",),
        minimum_points=1,
        test_policies=(
            CorrelationTestPolicy(
                "By number",
                ProfileTestSelector(exact=(101,)),
                "mean_delta",
                GuardBandProfile(kind="distribution_sigma"),
            ),
            CorrelationTestPolicy(
                "By name",
                ProfileTestSelector(name_contains=("Leakage",)),
                "median_offset",
                GuardBandProfile(kind="shifted_upper_limit"),
            ),
        ),
    )

    with pytest.raises(ValueError, match="Multiple test-set policies"):
        correlate_frame(frame, profile)


def test_minimum_point_error_identifies_over_grouping_dimension() -> None:
    rows = [
        {"DUT Nr": dut, "Test Number": 101, "Temperature": temperature, "reference": dut + 0.1, "ate": dut}
        for dut in range(1, 12)
        for temperature in (135, 135, 25, 25, -40)
    ]
    profile = CorrelationProfile(
        name="over-grouped",
        strategy="median_offset",
        reference_column="reference",
        candidate_column="ate",
        group_by=("DUT Nr", "Test Number", "Temperature"),
        minimum_points=5,
    )

    with pytest.raises(ValueError) as captured:
        correlate_frame(pd.DataFrame(rows), profile)

    message = str(captured.value)
    assert "55 valid Lab/CV-to-ATE value pairs formed 33 groups" in message
    assert "largest group contains 2 points" in message
    assert "remove 'DUT Nr'" in message
    assert "3 groups pass (largest group: 22)" in message
    assert "keep device identifiers such as DUT Nr in Detail key columns" in message


def test_empty_numeric_pair_error_reports_each_value_column() -> None:
    frame = pd.DataFrame({"group": ["A", "A"], "reference": ["", "N/A"], "ate": [1.0, 2.0]})
    profile = CorrelationProfile(
        name="invalid values",
        strategy="median_offset",
        reference_column="reference",
        candidate_column="ate",
        group_by=("group",),
        minimum_points=1,
    )

    with pytest.raises(ValueError, match="Numeric rows: 0 reference, 2 ATE, from 2 input rows"):
        correlate_frame(frame, profile)


def test_plot_title_reports_one_based_sample_count(tmp_path: Path) -> None:
    frame = pd.DataFrame({
        "DUT Nr": range(1, 6),
        "group": ["A"] * 5,
        "reference": [2, 3, 4, 5, 6],
        "ate": [1, 2, 3, 4, 5],
    })
    profile = CorrelationProfile(
        name="plot samples",
        strategy="mean_delta",
        reference_column="reference",
        candidate_column="ate",
        group_by=("group",),
        detail_key_columns=("DUT Nr",),
        minimum_points=5,
    )

    count = write_plots(correlate_frame(frame, profile), profile, tmp_path)

    assert count == 2
    fe_folder = tmp_path / f"{tmp_path.name}_FE"
    filenames = sorted(path.name for path in fe_folder.glob("*.png"))
    assert len(filenames) == 2
    assert all(filename.startswith("G0000_Samples_5_plot_samples_group_A_N_5_") for filename in filenames)
    assert filenames[0].endswith("__models.png")
    assert filenames[1].endswith("__series.png")


def test_correlated_series_omits_original_limits_and_autofits_to_models_and_new_limits(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from matplotlib.axes import Axes

    frame = pd.DataFrame({
        "group": ["A"] * 5,
        "reference": [9.8, 10.1, 10.4, 10.7, 11.0],
        "ate": [9.0, 9.3, 9.6, 9.9, 10.2],
        "low": [-1000.0] * 5,
        "high": [1000.0] * 5,
    })
    profile = CorrelationProfile(
        name="limit visibility",
        strategy="Mean_Deltas",
        reference_column="reference",
        candidate_column="ate",
        group_by=("group",),
        minimum_points=5,
        lower_limit_column="low",
        upper_limit_column="high",
        guard_band=GuardBandProfile(kind="distribution_sigma", sigma_multiplier=2.0),
    )
    horizontal_labels: list[str] = []
    corrected_limits: list[tuple[float, float]] = []
    original_axhline = Axes.axhline
    original_set_ylim = Axes.set_ylim

    def recording_axhline(self: Axes, *args: object, **kwargs: object) -> object:
        horizontal_labels.append(str(kwargs.get("label", "")))
        return original_axhline(self, *args, **kwargs)

    def recording_set_ylim(self: Axes, *args: object, **kwargs: object) -> object:
        result = original_set_ylim(self, *args, **kwargs)
        if self.get_title() == "Correlated series and limits":
            corrected_limits.append(tuple(float(value) for value in self.get_ylim()))
        return result

    monkeypatch.setattr(Axes, "axhline", recording_axhline)
    monkeypatch.setattr(Axes, "set_ylim", recording_set_ylim)

    write_plots(correlate_frame(frame, profile), profile, tmp_path)

    assert not any("Original" in label for label in horizontal_labels)
    assert not any("REQ_" in label for label in horizontal_labels)
    assert any("New LTL" in label for label in horizontal_labels)
    assert any("New UTL" in label for label in horizontal_labels)
    assert not any("Worst-case UTL" in label for label in horizontal_labels)
    assert corrected_limits
    low, high = corrected_limits[-1]
    assert low > 0.0
    assert high < 20.0


def test_test_set_pools_any_dimensions_into_one_shared_factor_and_guard_band(tmp_path: Path) -> None:
    frame = pd.DataFrame([
        {
            "DUT Nr": dut,
            "Test Number": 1000 + channel,
            "Frequency": 81,
            "Supply Corner": "VMIN",
            "Channel": channel,
            "Digital Control": 255,
            "Insertion": "S1",
            "Temperature": 135,
            "Test Name": f"TXPA_81FwLu255Tx{channel}_D095",
            "reference": 10.0 + channel + dut / 100,
            "ate": 9.5 + channel + dut / 100,
            "low": 9.0,
            "high": 16.0,
            "Unit": "dBm",
        }
        for dut in range(1, 12)
        for channel in range(1, 9)
    ])
    profile = CorrelationProfile(
        name="pooled TXPA",
        strategy="median_offset",
        reference_column="reference",
        candidate_column="ate",
        group_by=(
            "Test Number", "Frequency", "Supply Corner", "Channel", "Digital Control", "Insertion", "Temperature",
        ),
        minimum_points=5,
        lower_limit_column="low",
        upper_limit_column="high",
        unit_column="Unit",
        detail_key_columns=("DUT Nr",),
        test_policies=(CorrelationTestPolicy(
            "8-channel pool",
            ProfileTestSelector(ranges=((1001, 1008),)),
            "median_offset",
            GuardBandProfile(kind="distribution_sigma", sigma_multiplier=6),
            pooled_columns=("Test Number", "Channel"),
        ),),
    )

    result = correlate_frame(frame, profile)

    assert len(result.summary) == 1
    summary = result.summary.iloc[0]
    assert summary["Count"] == 88
    assert summary["Test Number"] == "MERGED"
    assert summary["Channel"] == "MERGED"
    assert summary["PooledParameters"] == "Test Number, Channel"
    assert summary["Merged Test Number Count"] == 8
    assert summary["Merged Channel Count"] == 8
    assert summary["Merged Test Number"] == "1001; 1002; 1003; 1004; 1005; 1006; 1007; 1008"
    assert summary["Merged Channel"] == "1; 2; 3; 4; 5; 6; 7; 8"
    assert summary["CorrelationFactor"] == pytest.approx(0.5)
    assert result.details["CorrelationFactor"].nunique() == 1
    assert result.details["AdjustedLowerLimit"].nunique() == 1
    assert result.details["AdjustedUpperLimit"].nunique() == 1
    assert set(result.details["Test Number"]) == set(range(1001, 1009))

    report = tmp_path / "pooled-report.xlsx"
    write_excel_report(result, profile, report)
    factors = pd.read_excel(report, sheet_name="Correlation_Factors")
    guards = pd.read_excel(report, sheet_name="Guard_Bands")
    for output in (factors, guards):
        assert output.loc[0, "Count"] == 88
        assert output.loc[0, "PooledParameters"] == "Test Number, Channel"
        assert output.loc[0, "Merged Test Number Count"] == 8
        assert output.loc[0, "Merged Channel Count"] == 8

    for sheet_name in ("Correlation_Factors", "Guard_Bands", "Correlation_Summary", "Correlated_Data"):
        columns = list(pd.read_excel(report, sheet_name=sheet_name, nrows=0).columns)
        assert columns[columns.index("Test Number") + 1] == "Test Name"

    plots = tmp_path / "plots"
    assert write_plots(result, profile, plots) == 2
    filenames = sorted(path.name for path in (plots / "plots_FE").glob("*.png"))
    assert len(filenames) == 2
    assert all(filename.startswith("G0000_Samples_88_") for filename in filenames)


def test_all_four_models_are_calculated_and_physics_can_be_primary(tmp_path: Path) -> None:
    kf = [0.0, 1.0, 2.0, 3.0, 10.0]
    ate = [20.0 + 3.0 * value for value in kf]
    cv = [ate_value - (2.0 * kf_value + 1.0) for ate_value, kf_value in zip(ate, kf)]
    frame = pd.DataFrame({
        "group": ["A"] * 5,
        "reference": cv,
        "ate": ate,
        "Kf": kf,
        "low": [0.0] * 5,
        "high": [100.0] * 5,
    })
    profile = CorrelationProfile(
        name="four models",
        strategy="Physics-based",
        reference_column="reference",
        candidate_column="ate",
        group_by=("group",),
        minimum_points=5,
        lower_limit_column="low",
        upper_limit_column="high",
        guard_band=GuardBandProfile(kind="Max_residuals", requirement_min=0.0, requirement_max=100.0),
        covariate=CovariateProfile("Kf source", ("group",), "Kf"),
    )

    result = correlate_frame(frame, profile)
    summary = result.summary.iloc[0]

    assert summary["LinearSlope"] == pytest.approx(1 / 3)
    assert summary["LinearIntercept"] == pytest.approx(37 / 3)
    assert summary["LinearR2"] == pytest.approx(1.0)
    assert summary["MeanDelta"] == pytest.approx(-7.4)
    assert pd.notna(summary["MeanDeltasR2"])
    assert summary["MedianDelta"] == pytest.approx(-5.0)
    assert summary["PhysicsAlpha"] == pytest.approx(2.0)
    assert summary["PhysicsBeta"] == pytest.approx(1.0)
    assert summary["PhysicsR2"] == pytest.approx(1.0)
    assert summary["PhysicsResidualStd"] == pytest.approx(0.0)
    assert summary["CorrelationStrategy"] == "Physics-based"
    assert pd.isna(summary["CorrelationFactor"])
    assert summary["CorrelationFactorA"] == pytest.approx(2.0)
    assert summary["CorrelationFactorB"] == pytest.approx(1.0)
    assert summary["AdjustedLowerLimit"] == pytest.approx(0.0)
    assert summary["AdjustedUpperLimit"] == pytest.approx(100.0)
    assert result.details["MeanDeltasCorrectedCandidate"].tolist() == pytest.approx(
        [value - 7.4 for value in ate]
    )
    assert result.details["PhysicsCorrectedCandidate"].tolist() == pytest.approx(cv)
    assert result.details["CorrectedCandidate"].tolist() == pytest.approx(cv)
    assert {
        "LinearCorrectedCandidate",
        "MeanDeltasCorrectedCandidate",
        "MedianDeltasCorrectedCandidate",
        "PhysicsCorrectedCandidate",
        "LinearResidual",
        "MeanDeltasResidual",
        "MedianDeltasResidual",
        "PhysicsResidual",
    }.issubset(result.details.columns)

    report = tmp_path / "physics-primary.xlsx"
    write_excel_report(result, profile, report)
    factors = pd.read_excel(report, sheet_name="Correlation_Factors")
    assert "CorrelationFactor" not in factors.columns
    assert factors.loc[0, "CorrelationFactorA"] == pytest.approx(2.0)
    assert factors.loc[0, "CorrelationFactorB"] == pytest.approx(1.0)
    assert not {
        "MeanDelta", "MedianDelta", "LinearSlope", "LinearIntercept", "PhysicsAlpha", "PhysicsBeta",
    }.intersection(factors.columns)


def test_linear_strategy_uses_ols_slope_and_intercept_as_factors() -> None:
    frame = pd.DataFrame({
        "group": ["A"] * 5,
        "reference": [3.0, 5.0, 7.0, 9.0, 11.0],
        "ate": [1.0, 2.0, 3.0, 4.0, 5.0],
        "low": [0.0] * 5,
        "high": [20.0] * 5,
    })
    profile = CorrelationProfile(
        name="OLS linear",
        strategy="Linear",
        reference_column="reference",
        candidate_column="ate",
        group_by=("group",),
        minimum_points=5,
        lower_limit_column="low",
        upper_limit_column="high",
        guard_band=GuardBandProfile(kind="Max_residuals", requirement_min=0.0, requirement_max=20.0),
    )

    result = correlate_frame(frame, profile)
    summary = result.summary.iloc[0]

    assert summary["LinearSlope"] == pytest.approx(2.0)
    assert summary["LinearIntercept"] == pytest.approx(1.0)
    assert summary["CorrelationFactorA"] == pytest.approx(2.0)
    assert summary["CorrelationFactorB"] == pytest.approx(1.0)
    assert pd.isna(summary["CorrelationFactor"])
    assert summary["LinearR2"] == pytest.approx(1.0)
    assert summary["MaxAbsResidual"] == pytest.approx(0.0)
    assert result.details["CorrectedCandidate"].tolist() == pytest.approx(frame["reference"].tolist())


def test_covariate_lookup_joins_kf_and_rejects_conflicting_values() -> None:
    frame = pd.DataFrame({
        "DUT Nr": [1, 2],
        "Temperature": [25.0, -40.0],
        "reference": [10.0, 11.0],
        "ate": [9.0, 10.0],
    })
    lookup = pd.DataFrame({
        "DUT Nr": [1.0, 1.0, 2.0],
        "Temperature": [25, 25, -40],
        "Kf measured": [0.25, 0.25, 0.5],
    })
    profile = CorrelationProfile(
        name="Kf lookup",
        strategy="Physics-based",
        reference_column="reference",
        candidate_column="ate",
        group_by=("Temperature",),
        minimum_points=1,
        covariate=CovariateProfile("Kf measured", ("DUT Nr", "Temperature"), "Kf"),
    )

    merged = attach_covariate(frame, lookup, profile)
    assert merged["Kf"].tolist() == pytest.approx([0.25, 0.5])

    conflicting = pd.concat([
        lookup,
        pd.DataFrame({"DUT Nr": [1], "Temperature": [25], "Kf measured": [0.75]}),
    ], ignore_index=True)
    with pytest.raises(ValueError, match="exactly one Kf value per DUT Nr, Temperature combination"):
        attach_covariate(frame, conflicting, profile)


def test_covariate_is_extracted_from_raw_test_rows_and_attached_to_targets() -> None:
    frame = pd.DataFrame({
        "DUT Nr": [1, 2, 1, 2],
        "Temperature": [25, 25, 25, 25],
        "Test Number": [52046, 52046, 101, 101],
        "Test Value": [0.25, 0.5, 10.0, 11.0],
    })
    profile = CorrelationProfile(
        name="raw Kf",
        strategy="Physics-based",
        reference_column="reference",
        candidate_column="ate",
        group_by=("Temperature",),
        covariate=CovariateProfile("Test Value", ("DUT Nr", "Temperature"), "Kf", 52046),
    )

    prepared = attach_covariate_from_test_rows(frame, profile)

    assert prepared["Test Number"].tolist() == [101, 101]
    assert prepared["Test Value"].tolist() == pytest.approx([10.0, 11.0])
    assert prepared["Kf"].tolist() == pytest.approx([0.25, 0.5])


def test_max_residuals_tightens_configured_requirement_limits() -> None:
    frame = pd.DataFrame({
        "group": ["A"] * 5,
        "reference": [10, 11, 12, 13, 20],
        "ate": [9, 10, 11, 12, 12],
        "low": [0] * 5,
        "high": [20] * 5,
    })
    profile = CorrelationProfile(
        name="maximum residual limits",
        strategy="Median_Deltas",
        reference_column="reference",
        candidate_column="ate",
        group_by=("group",),
        minimum_points=5,
        lower_limit_column="low",
        upper_limit_column="high",
        guard_band=GuardBandProfile(kind="Max_residuals", requirement_min=2.0, requirement_max=18.0),
    )

    summary = correlate_frame(frame, profile).summary.iloc[0]

    assert summary["MaxAbsResidual"] == pytest.approx(7.0)
    assert summary["RequirementMin"] == pytest.approx(2.0)
    assert summary["RequirementMax"] == pytest.approx(18.0)
    assert summary["OriginalLowerLimit"] == pytest.approx(0.0)
    assert summary["OriginalUpperLimit"] == pytest.approx(20.0)
    assert summary["AdjustedLowerLimit"] == pytest.approx(9.0)
    assert summary["AdjustedUpperLimit"] == pytest.approx(11.0)
    assert not summary["LimitWindowInvalid"]


def test_physics_primary_rejects_constant_kf_with_actionable_message() -> None:
    frame = pd.DataFrame({
        "group": ["A"] * 5,
        "reference": [10, 11, 12, 13, 14],
        "ate": [9, 10, 11, 12, 13],
        "Kf": [1.0] * 5,
    })
    profile = CorrelationProfile(
        name="constant Kf",
        strategy="Physics-based",
        reference_column="reference",
        candidate_column="ate",
        group_by=("group",),
        minimum_points=5,
        covariate=CovariateProfile("Kf source", ("group",), "Kf"),
    )

    with pytest.raises(ValueError, match="Kf has insufficient variation"):
        correlate_frame(frame, profile)


def test_plots_are_dispatched_to_fe_and_be_with_two_files_per_group(tmp_path: Path) -> None:
    frame = pd.DataFrame([
        {
            "DUT Nr": dut,
            "Test Number": 101,
            "Test Name": "Power",
            "Insertion": insertion,
            "Insertion Type": insertion_type,
            "reference": float(dut),
            "ate": float(dut) - 0.5,
        }
        for insertion, insertion_type in (("S1", "FE"), ("B1", "BE"))
        for dut in range(1, 6)
    ])
    profile = CorrelationProfile(
        name="insertion plots",
        strategy="Linear",
        reference_column="reference",
        candidate_column="ate",
        group_by=("Test Number", "Insertion"),
        detail_key_columns=("DUT Nr",),
        minimum_points=5,
    )

    count = write_plots(correlate_frame(frame, profile), profile, tmp_path)

    assert count == 4
    for insertion_type in ("FE", "BE"):
        files = list((tmp_path / f"{tmp_path.name}_{insertion_type}").glob("*.png"))
        assert len(files) == 2
        assert {path.name.rsplit("__", 1)[1] for path in files} == {"series.png", "models.png"}


def test_pooled_group_rejects_incompatible_limits() -> None:
    frame = pd.DataFrame({
        "Test Number": [101] * 5 + [102] * 5,
        "group": ["A"] * 10,
        "reference": range(10),
        "ate": range(10),
        "high": [10] * 5 + [20] * 5,
    })
    profile = CorrelationProfile(
        name="incompatible pooled limits",
        strategy="median_offset",
        reference_column="reference",
        candidate_column="ate",
        group_by=("Test Number", "group"),
        upper_limit_column="high",
        minimum_points=5,
        pooled_columns=("Test Number",),
    )

    with pytest.raises(ValueError, match="different values in limit column 'high'"):
        correlate_frame(frame, profile)


def test_test_sets_can_pool_different_parameters_without_cross_pooling() -> None:
    frame = pd.DataFrame([
        {
            "Test Number": test_number,
            "Channel": channel,
            "condition": "A",
            "reference": float(sample + 1),
            "ate": float(sample),
        }
        for test_number, channel in ((101, 1), (102, 1), (201, 1), (201, 2))
        for sample in range(5)
    ])
    profile = CorrelationProfile(
        name="independent pools",
        strategy="mean_delta",
        reference_column="reference",
        candidate_column="ate",
        group_by=("Test Number", "Channel", "condition"),
        minimum_points=5,
        test_policies=(
            CorrelationTestPolicy(
                "pool tests",
                ProfileTestSelector(ranges=((101, 102),)),
                "mean_delta",
                GuardBandProfile(kind="distribution_sigma"),
                pooled_columns=("Test Number",),
            ),
            CorrelationTestPolicy(
                "pool channels",
                ProfileTestSelector(exact=(201,)),
                "mean_delta",
                GuardBandProfile(kind="distribution_sigma"),
                pooled_columns=("Channel",),
            ),
        ),
    )

    result = correlate_frame(frame, profile)

    assert len(result.summary) == 2
    assert set(result.summary["TestSet"]) == {"pool tests", "pool channels"}
    assert set(result.summary["PooledParameters"]) == {"Test Number", "Channel"}
    assert set(result.summary["Count"]) == {10}
    assert result.details.groupby("TestSet").size().to_dict() == {"pool channels": 10, "pool tests": 10}
