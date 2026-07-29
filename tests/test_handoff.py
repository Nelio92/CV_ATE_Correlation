from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from cv_ate_correlation.handoff import (
    MANIFEST_SHEET,
    REPEAT_INDEX,
    REQUEST_ID,
    REQUEST_SHEET,
    create_measurement_request,
    import_measurement_results,
)
from cv_ate_correlation.models import CorrelationProfile


PROFILE = CorrelationProfile(
    name="Generic voltage campaign",
    strategy="median_offset",
    reference_column="CV_Value",
    candidate_column="ATE_Value",
    group_by=("Parameter", "Temperature"),
)


def _source() -> pd.DataFrame:
    return pd.DataFrame({
        "DUT": [1, 1, 2],
        "Parameter": ["VOUT", "VOUT", "VOUT"],
        "Temperature": [25, 25, 25],
        "Unit": ["V", "V", "V"],
        "Test Value": [1.00, 1.01, 0.99],
        "Low": [0.9, 0.9, 0.9],
        "High": [1.1, 1.1, 1.1],
    })


def _fill_returned(path: Path, values: list[float]) -> None:
    workbook = load_workbook(path)
    sheet = workbook[REQUEST_SHEET]
    reference_column = next(cell.column for cell in sheet[1] if cell.value == PROFILE.reference_column)
    for row, value in enumerate(values, start=2):
        sheet.cell(row=row, column=reference_column, value=value)
    workbook.save(path)


def test_unprotected_request_with_id_warning_round_trip_is_one_to_one(tmp_path: Path) -> None:
    request_path, manifest_path = tmp_path / "request.xlsx", tmp_path / "manifest.xlsx"
    request, manifest = create_measurement_request(_source(), PROFILE, request_path, manifest_path)

    assert list(request[REPEAT_INDEX]) == [1, 2, 1]
    assert not request[REQUEST_ID].is_unique
    assert not request.duplicated([REQUEST_ID, REPEAT_INDEX]).any()
    assert "Test Value" not in request.columns
    assert "Low" not in request.columns
    assert "High" not in request.columns
    assert list(manifest[PROFILE.candidate_column]) == [1.00, 1.01, 0.99]

    workbook = load_workbook(request_path)
    request_sheet = workbook[REQUEST_SHEET]
    assert not request_sheet.protection.sheet
    assert request_sheet.auto_filter.ref == request_sheet.dimensions
    assert workbook["_Metadata"].sheet_state == "veryHidden"
    request_id_column = next(cell.column for cell in request_sheet[1] if cell.value == REQUEST_ID)
    reference_column = next(cell.column for cell in request_sheet[1] if cell.value == PROFILE.reference_column)
    assert all(
        not request_sheet.cell(row=row, column=column).protection.locked
        for row in range(1, request_sheet.max_row + 1)
        for column in range(1, request_sheet.max_column + 1)
    )
    assert not list(request_sheet.data_validations.dataValidation)
    warning = request_sheet.cell(row=1, column=request_id_column).comment
    assert warning is not None
    assert "Do not modify any Measurement_Request_ID value" in warning.text
    assert not request_sheet.cell(row=2, column=reference_column).protection.locked
    assert not request_sheet.cell(row=2, column=request_id_column).protection.locked

    _fill_returned(request_path, [1.02, 1.03, 1.00])
    aligned = import_measurement_results(request_path, manifest_path, PROFILE)
    assert list(aligned[PROFILE.reference_column]) == [1.02, 1.03, 1.00]
    assert list(aligned[PROFILE.candidate_column]) == [1.00, 1.01, 0.99]


def test_duplicate_returned_request_key_is_rejected(tmp_path: Path) -> None:
    request_path, manifest_path = tmp_path / "request.xlsx", tmp_path / "manifest.xlsx"
    create_measurement_request(_source(), PROFILE, request_path, manifest_path)
    _fill_returned(request_path, [1.02, 1.03, 1.00])

    workbook = load_workbook(request_path)
    sheet = workbook[REQUEST_SHEET]
    sheet.cell(row=3, column=1, value=sheet.cell(row=2, column=1).value)
    sheet.cell(row=3, column=2, value=sheet.cell(row=2, column=2).value)
    workbook.save(request_path)

    with pytest.raises(ValueError, match="duplicate measurement request keys"):
        import_measurement_results(request_path, manifest_path, PROFILE)


def test_modified_request_id_is_rejected_by_strict_import(tmp_path: Path) -> None:
    request_path, manifest_path = tmp_path / "request.xlsx", tmp_path / "manifest.xlsx"
    create_measurement_request(_source(), PROFILE, request_path, manifest_path)

    workbook = load_workbook(request_path)
    sheet = workbook[REQUEST_SHEET]
    sheet.cell(row=2, column=1, value="MR-MODIFIED")
    workbook.save(request_path)

    with pytest.raises(ValueError, match="Request coverage mismatch"):
        import_measurement_results(request_path, manifest_path, PROFILE)


def test_request_creation_does_not_rescan_sheet_bounds_for_every_row(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    max_row_getter = Worksheet.max_row.fget
    max_column_getter = Worksheet.max_column.fget
    assert max_row_getter is not None
    assert max_column_getter is not None
    reads = {"row": 0, "column": 0}

    def counted_max_row(worksheet: Worksheet) -> int:
        reads["row"] += 1
        return max_row_getter(worksheet)

    def counted_max_column(worksheet: Worksheet) -> int:
        reads["column"] += 1
        return max_column_getter(worksheet)

    monkeypatch.setattr(Worksheet, "max_row", property(counted_max_row))
    monkeypatch.setattr(Worksheet, "max_column", property(counted_max_column))
    source = pd.concat([_source()] * 50, ignore_index=True)

    create_measurement_request(source, PROFILE, tmp_path / "request.xlsx", tmp_path / "manifest.xlsx")

    assert reads["row"] < 20
    assert reads["column"] < 20


def test_manifest_contains_internal_sheet(tmp_path: Path) -> None:
    request_path, manifest_path = tmp_path / "request.xlsx", tmp_path / "manifest.xlsx"
    create_measurement_request(_source(), PROFILE, request_path, manifest_path)
    assert MANIFEST_SHEET in pd.ExcelFile(manifest_path).sheet_names
