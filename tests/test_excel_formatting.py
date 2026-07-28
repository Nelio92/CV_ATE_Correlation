from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from cv_ate_correlation.excel import ACCENT_1_BLUE, write_dataframe_workbook
from cv_ate_correlation.handoff import create_measurement_request
from cv_ate_correlation.models import CorrelationProfile


def assert_sheet_is_formatted(path: Path, sheet_name: str) -> None:
    worksheet = load_workbook(path)[sheet_name]
    assert worksheet.auto_filter.ref == worksheet.dimensions
    assert worksheet.freeze_panes == "A2"
    for cell in worksheet[1]:
        assert cell.fill.fill_type == "solid"
        assert cell.fill.fgColor.rgb == f"00{ACCENT_1_BLUE}"
        assert cell.font.bold
        assert cell.font.color is not None
        assert cell.font.color.rgb == "00FFFFFF"
    for column_index in range(1, worksheet.max_column + 1):
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        assert worksheet.column_dimensions[column_letter].width >= 10


def test_standard_writer_formats_every_sheet_and_autofits_columns(tmp_path: Path) -> None:
    output = tmp_path / "formatted.xlsx"
    write_dataframe_workbook(output, {
        "First": pd.DataFrame({"Long header name": ["short"], "Value column": [123]}),
        "Second": pd.DataFrame({"Another long header": ["longer cell value"], "Status value": ["OK"]}),
    })

    assert_sheet_is_formatted(output, "First")
    assert_sheet_is_formatted(output, "Second")


def test_request_and_internal_manifest_receive_standard_formatting(tmp_path: Path) -> None:
    request_path = tmp_path / "request.xlsx"
    manifest_path = tmp_path / "manifest.xlsx"
    profile = CorrelationProfile(
        name="Formatting campaign",
        strategy="median_offset",
        reference_column="Lab Value",
        candidate_column="ATE Value",
        group_by=("Test Number",),
    )
    source = pd.DataFrame({
        "Test Number": [101, 102],
        "Test Name": ["A long descriptive test name", "Another test"],
        "Test Value": [1.0, 2.0],
    })

    create_measurement_request(source, profile, request_path, manifest_path)

    for sheet_name in load_workbook(request_path).sheetnames:
        assert_sheet_is_formatted(request_path, sheet_name)
    for sheet_name in load_workbook(manifest_path).sheetnames:
        assert_sheet_is_formatted(manifest_path, sheet_name)
