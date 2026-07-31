"""Canonical Section 2 chip-manifest template and distribution helpers."""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

CHIP_MANIFEST_FILENAME = "CorreLaTE_Chips_Manifest.xlsx"
CHIP_MANIFEST_SHEET = "Chip_Manifest"
CHIP_MANIFEST_COLUMNS = ("DUT Nr", "Wafer", "X", "Y", "DoE split")
CHIP_MANIFEST_INPUT_LAST_ROW = 5000

_ASSET_DIRECTORY = Path(__file__).resolve().with_name("assets")
_BLUE = "1F4E78"
_BLUE_LIGHT = "D9EAF7"
_GREEN = "278F6A"
_GREEN_LIGHT = "E2F0D9"
_GOLD = "E2A93B"
_GOLD_LIGHT = "FFF2CC"
_RED = "C00000"
_RED_LIGHT = "F4CCCC"
_WHITE = "FFFFFF"
_TEXT = "263746"
_BORDER = Side(style="thin", color="B8C4CE")


def chip_manifest_template_path() -> Path:
    """Return the packaged blank chip-manifest workbook."""
    return _ASSET_DIRECTORY / CHIP_MANIFEST_FILENAME


def save_chip_manifest_template(destination: str | Path) -> Path:
    """Copy the packaged template to a user-selected location."""
    source = chip_manifest_template_path()
    if not source.is_file():
        raise FileNotFoundError(f"Packaged chip-manifest template is missing: {source}")
    target = Path(destination)
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(source, target)
    return target


def _add_validation(
    worksheet,
    *,
    validation_type: str,
    formula: str,
    cells: str,
    prompt_title: str,
    prompt: str,
    error: str,
    error_style: str = "stop",
    formula2: str | None = None,
) -> None:
    validation = DataValidation(
        type=validation_type,
        formula1=formula,
        formula2=formula2,
        allow_blank=False,
        showInputMessage=True,
        promptTitle=prompt_title,
        prompt=prompt,
        showErrorMessage=True,
        errorTitle="Invalid chip-manifest value",
        error=error,
        errorStyle=error_style,
    )
    worksheet.add_data_validation(validation)
    validation.add(cells)


def create_chip_manifest_template(destination: str | Path) -> Path:
    """Create the branded, one-sheet Section 2 chip-manifest template."""
    target = Path(destination)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.creator = "CorreLaTE"
    workbook.title = "CorreLaTE Section 2 Chip Manifest"
    workbook.subject = "Required chip identity and process-corner input for ATE extraction"
    workbook.description = (
        "Populate all five columns with one unique physical chip per row before running Section 2."
    )
    worksheet = workbook.active
    worksheet.title = CHIP_MANIFEST_SHEET
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.print_area = "A1:K22"
    worksheet.sheet_view.selection[0].activeCell = "A2"
    worksheet.sheet_view.selection[0].sqref = "A2"

    header_fill = PatternFill("solid", fgColor=_BLUE)
    input_fill = PatternFill("solid", fgColor=_GOLD_LIGHT)
    for column, value in enumerate(CHIP_MANIFEST_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=column, value=value)
        cell.fill = header_fill
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=_BORDER, bottom=_BORDER, left=_BORDER, right=_BORDER)
        input_cell = worksheet.cell(row=2, column=column)
        input_cell.fill = input_fill
        input_cell.border = Border(top=_BORDER, bottom=_BORDER, left=_BORDER, right=_BORDER)

    comments = {
        "A1": (
            "REQUIRED. Unique device number used by the CV team to identify the physical chip and align "
            "the returned CV result with TE data. Enter one DUT number per chip; duplicates are rejected."
        ),
        "B1": (
            "REQUIRED. Wafer identifier exactly matching the raw TE export. Numeric and text wafer IDs are supported."
        ),
        "C1": "REQUIRED. Integer X coordinate of the die in the raw TE export.",
        "D1": "REQUIRED. Integer Y coordinate of the die in the raw TE export.",
        "E1": (
            "REQUIRED. Process-corner or DoE split used to separate correlation plots and interpret process variation. "
            "Common values are TT, SS, and FF; project-specific split labels are also accepted."
        ),
    }
    for coordinate, text in comments.items():
        worksheet[coordinate].comment = Comment(text, "CorreLaTE")

    table = Table(displayName="CorreLaTEChipManifest", ref="A1:E2")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

    last_row = CHIP_MANIFEST_INPUT_LAST_ROW
    _add_validation(
        worksheet,
        validation_type="whole",
        formula="1",
        formula2="2147483647",
        cells=f"A2:A{last_row}",
        prompt_title="Required: DUT Nr",
        prompt="Enter the unique positive DUT number used by the CV team.",
        error="DUT Nr must be a positive whole number.",
    )
    _add_validation(
        worksheet,
        validation_type="custom",
        formula='LEN(TRIM(B2&""))>0',
        cells=f"B2:B{last_row}",
        prompt_title="Required: Wafer",
        prompt="Enter the wafer ID exactly as represented in the raw TE data.",
        error="Wafer cannot be blank.",
    )
    for column, label in (("C", "X"), ("D", "Y")):
        _add_validation(
            worksheet,
            validation_type="whole",
            formula="-32768",
            formula2="32767",
            cells=f"{column}2:{column}{last_row}",
            prompt_title=f"Required: {label} coordinate",
            prompt=f"Enter the integer die {label} coordinate from the raw TE data.",
            error=f"{label} must be a whole-number die coordinate.",
        )
    _add_validation(
        worksheet,
        validation_type="list",
        formula='"TT,SS,FF"',
        cells=f"E2:E{last_row}",
        prompt_title="Required: DoE split",
        prompt="Choose TT/SS/FF or type the applicable project-specific process-corner label.",
        error="Use TT, SS, FF, or confirm the project-specific process-corner label.",
        error_style="warning",
    )

    missing_fill = PatternFill("solid", fgColor=_RED_LIGHT)
    duplicate_fill = PatternFill("solid", fgColor="F8CBAD")
    for column in "ABCDE":
        worksheet.conditional_formatting.add(
            f"{column}2:{column}{last_row}",
            FormulaRule(
                formula=[f'AND(COUNTA($A2:$E2)>0,{column}2="")'],
                fill=missing_fill,
            ),
        )
    worksheet.conditional_formatting.add(
        f"A2:E{last_row}",
        FormulaRule(
            formula=[f'AND($A2<>"",COUNTIF($A$2:$A${last_row},$A2)>1)'],
            fill=duplicate_fill,
        ),
    )
    worksheet.conditional_formatting.add(
        f"A2:E{last_row}",
        FormulaRule(
            formula=[
                f'AND(COUNTA($B2:$D2)=3,COUNTIFS($B$2:$B${last_row},$B2,'
                f'$C$2:$C${last_row},$C2,$D$2:$D${last_row},$D2)>1)'
            ],
            fill=duplicate_fill,
        ),
    )

    worksheet.merge_cells("G1:J2")
    title = worksheet["G1"]
    title.value = "CorreLaTE\nChip Manifest — Section 2"
    title.fill = PatternFill("solid", fgColor=_BLUE)
    title.font = Font(color=_WHITE, size=18, bold=True)
    title.alignment = Alignment(vertical="center", wrap_text=True)
    for row in worksheet["G1:J2"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=_BLUE)

    logo_path = _ASSET_DIRECTORY / "correlate-signal-bloom-64.png"
    if logo_path.is_file():
        logo = Image(str(logo_path))
        logo.width = 56
        logo.height = 56
        worksheet.add_image(logo, "K1")

    worksheet.merge_cells("G4:K4")
    notice = worksheet["G4"]
    notice.value = "All five columns are REQUIRED. One completed row represents one physical chip."
    notice.fill = PatternFill("solid", fgColor=_GOLD)
    notice.font = Font(color="573B00", bold=True)
    notice.alignment = Alignment(horizontal="center", vertical="center")

    guide_headers = ("Column", "Required", "Format", "What to enter", "Why it matters")
    for column, value in enumerate(guide_headers, start=7):
        cell = worksheet.cell(row=6, column=column, value=value)
        cell.fill = PatternFill("solid", fgColor=_GREEN)
        cell.font = Font(color=_WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=_BORDER, bottom=_BORDER, left=_BORDER, right=_BORDER)

    guidance = (
        ("DUT Nr", "YES", "Whole number", "The CV team's chip number, e.g. 1773.", "Primary TE–CV alignment identity."),
        ("Wafer", "YES", "ID / number", "Wafer ID exactly as in raw TE data, e.g. 2.", "Locates the chip in raw exports."),
        ("X", "YES", "Integer", "Die X coordinate, e.g. 31.", "Matches the physical die coordinate."),
        ("Y", "YES", "Integer", "Die Y coordinate, e.g. 5.", "Matches the physical die coordinate."),
        ("DoE split", "YES", "Text", "Process corner, e.g. TT, SS, FF, or project label.", "Enables split plots and process-variation review."),
    )
    for row_index, values in enumerate(guidance, start=7):
        for column, value in enumerate(values, start=7):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            cell.fill = PatternFill("solid", fgColor=_GREEN_LIGHT if row_index % 2 else "F4FAF7")
            cell.font = Font(color=_TEXT, bold=column in {7, 8})
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(top=_BORDER, bottom=_BORDER, left=_BORDER, right=_BORDER)

    worksheet.merge_cells("G13:K13")
    checklist_header = worksheet["G13"]
    checklist_header.value = "Before running Section 2 · Extract TE"
    checklist_header.fill = PatternFill("solid", fgColor=_BLUE_LIGHT)
    checklist_header.font = Font(color=_BLUE, bold=True, size=12)
    checklist_header.alignment = Alignment(vertical="center")
    checklist = (
        "1. Enter one chip per row in columns A–E; the Excel table expands automatically.",
        "2. Complete every field. Blank DUT Nr or DoE split values are rejected, not silently ignored.",
        "3. Keep each DUT Nr and each Wafer/X/Y combination unique.",
        "4. Keep the worksheet and column headers unchanged; save the workbook as .xlsx.",
        "5. In CorreLaTE Section 2, select this workbook as the Chip manifest input.",
    )
    for row_index, value in enumerate(checklist, start=14):
        worksheet.merge_cells(start_row=row_index, start_column=7, end_row=row_index, end_column=11)
        cell = worksheet.cell(row=row_index, column=7, value=value)
        cell.font = Font(color=_TEXT)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    worksheet.merge_cells("G20:K20")
    legend_header = worksheet["G20"]
    legend_header.value = "Visual checks"
    legend_header.fill = PatternFill("solid", fgColor=_BLUE_LIGHT)
    legend_header.font = Font(color=_BLUE, bold=True)
    worksheet["G21"].fill = input_fill
    worksheet["G21"].border = Border(top=_BORDER, bottom=_BORDER, left=_BORDER, right=_BORDER)
    worksheet.merge_cells("H21:K21")
    worksheet["H21"] = "Yellow cells are ready for user input."
    worksheet["G22"].fill = missing_fill
    worksheet["G22"].border = Border(top=_BORDER, bottom=_BORDER, left=_BORDER, right=_BORDER)
    worksheet.merge_cells("H22:K22")
    worksheet["H22"] = "Red/orange highlighting identifies missing or duplicate required identities."

    widths = {
        "A": 14,
        "B": 13,
        "C": 10,
        "D": 10,
        "E": 18,
        "F": 3,
        "G": 16,
        "H": 12,
        "I": 16,
        "J": 32,
        "K": 34,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.row_dimensions[1].height = 30
    worksheet.row_dimensions[2].height = 30
    worksheet.row_dimensions[4].height = 24
    worksheet.row_dimensions[6].height = 28
    for row_index in range(7, 12):
        worksheet.row_dimensions[row_index].height = 42
    for row_index in range(14, 19):
        worksheet.row_dimensions[row_index].height = 24

    workbook.save(target)
    return target
