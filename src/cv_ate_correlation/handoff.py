"""Editable measurement-request handoff with strict one-to-one result import."""

from __future__ import annotations

import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font, PatternFill, Protection

from .excel import ACCENT_1_BLUE, format_worksheet
from .models import CorrelationProfile

REQUEST_SHEET = "Measurement_Request"
MANIFEST_SHEET = "ATE_Manifest"
METADATA_SHEET = "_Metadata"
REQUEST_ID = "Measurement_Request_ID"
REPEAT_INDEX = "Repeat_Index"
SCHEMA_VERSION = "1"


def _canonical_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if hasattr(value, "item"):
        value = value.item()
    return value if isinstance(value, (str, int, float, bool)) else str(value)


def _request_id(profile_name: str, columns: list[str], row: pd.Series) -> str:
    payload = {
        "profile": profile_name,
        "conditions": {column: _canonical_value(row[column]) for column in columns},
    }
    digest = hashlib.sha256(json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")).hexdigest()
    return f"MR-{digest[:20].upper()}"


def _metadata(profile: CorrelationProfile, row_count: int) -> pd.DataFrame:
    return pd.DataFrame({
        "Key": ["Schema_Version", "Profile", "Reference_Column", "Candidate_Column", "Row_Count", "Generated_UTC"],
        "Value": [
            SCHEMA_VERSION,
            profile.name,
            profile.reference_column,
            profile.candidate_column,
            row_count,
            datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        ],
    })


def _prepare_request(workbook: Workbook, reference_column: str) -> None:
    for worksheet in workbook.worksheets:
        format_worksheet(worksheet)
    request = workbook[REQUEST_SHEET]
    request.freeze_panes = "A2"
    request.auto_filter.ref = request.dimensions
    request.sheet_view.showGridLines = False
    request.protection.sheet = False
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    identifier_fill = PatternFill("solid", fgColor="E7E6E6")
    for cell in request[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor=ACCENT_1_BLUE)
    request_id_index = next(cell.column for cell in request[1] if cell.value == REQUEST_ID)
    reference_index = next(
        cell.column for cell in request[1] if cell.value == reference_column
    )
    max_row = request.max_row
    max_column = request.max_column
    unlocked = Protection(locked=False)
    for row in range(1, max_row + 1):
        for column in range(1, max_column + 1):
            request.cell(row=row, column=column).protection = unlocked
    for row in range(2, max_row + 1):
        request.cell(row=row, column=request_id_index).fill = identifier_fill
        request.cell(row=row, column=reference_index).fill = input_fill
    request.cell(row=1, column=request_id_index).comment = Comment(
        "WARNING: Do not modify any Measurement_Request_ID value. These IDs are required to align returned CV "
        "measurements with the internal ATE manifest. The worksheet is intentionally unprotected so all cells can "
        "be selected, filtered, grouped, and sorted.",
        "CorreLaTE",
    )
    workbook[METADATA_SHEET].sheet_state = "veryHidden"


def create_measurement_request(
    frame: pd.DataFrame,
    profile: CorrelationProfile,
    request_path: Path,
    manifest_path: Path,
    *,
    candidate_value_column: str = "Test Value",
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Create an editable CV request and a separate TE-only ATE manifest."""
    if candidate_value_column not in frame.columns:
        raise ValueError(f"Input is missing candidate value column '{candidate_value_column}'")
    if frame.empty:
        raise ValueError("Cannot create a measurement request from an empty input")

    excluded = {candidate_value_column, profile.reference_column, profile.candidate_column, "Low", "High"}
    descriptor_columns = [column for column in frame.columns if column not in excluded]
    if not descriptor_columns:
        raise ValueError("No measurement descriptors remain after removing value columns")

    request = frame[descriptor_columns].copy()
    request.insert(0, REQUEST_ID, [
        _request_id(profile.name, descriptor_columns, row) for _, row in request.iterrows()
    ])
    request.insert(1, REPEAT_INDEX, request.groupby(REQUEST_ID, sort=False).cumcount() + 1)
    request[profile.reference_column] = pd.NA

    manifest = frame.copy()
    manifest.insert(0, REQUEST_ID, request[REQUEST_ID].to_numpy())
    manifest.insert(1, REPEAT_INDEX, request[REPEAT_INDEX].to_numpy())
    manifest[profile.candidate_column] = pd.to_numeric(manifest[candidate_value_column], errors="coerce")
    if manifest[profile.candidate_column].isna().any():
        count = int(manifest[profile.candidate_column].isna().sum())
        raise ValueError(f"Candidate value column contains {count} non-numeric or blank values")

    request_path = Path(request_path)
    manifest_path = Path(manifest_path)
    request_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    metadata = _metadata(profile, len(request))
    instructions = pd.DataFrame({"Instructions": [
        f"WARNING: Do not modify values in the '{REQUEST_ID}' column; they are required for one-to-one alignment.",
        f"Enter one numeric result in the yellow '{profile.reference_column}' column for every row.",
        f"Do not modify '{REPEAT_INDEX}'. It is also validated as part of the one-to-one key.",
        "Do not add, remove, or duplicate request rows.",
        "The worksheet is unprotected; select all cells and use header filters or Data > Sort for hierarchical grouping.",
        "Return this workbook to TE; ATE values are retained only in the separate internal manifest.",
    ]})
    with pd.ExcelWriter(request_path, engine="openpyxl") as writer:
        request.to_excel(writer, index=False, sheet_name=REQUEST_SHEET)
        instructions.to_excel(writer, index=False, sheet_name="Instructions")
        metadata.to_excel(writer, index=False, sheet_name=METADATA_SHEET)
        _prepare_request(writer.book, profile.reference_column)
    with pd.ExcelWriter(manifest_path, engine="openpyxl") as writer:
        manifest.to_excel(writer, index=False, sheet_name=MANIFEST_SHEET)
        metadata.to_excel(writer, index=False, sheet_name=METADATA_SHEET)
        for worksheet in writer.book.worksheets:
            format_worksheet(worksheet)
    return request, manifest


def _read_metadata(path: Path) -> dict[str, str]:
    try:
        metadata = pd.read_excel(path, sheet_name=METADATA_SHEET)
    except ValueError as error:
        raise ValueError(f"Workbook '{path}' has no {METADATA_SHEET} sheet") from error
    if not {"Key", "Value"}.issubset(metadata.columns):
        raise ValueError(f"Workbook '{path}' has invalid handoff metadata")
    return dict(zip(metadata["Key"].astype(str), metadata["Value"].astype(str)))


def import_measurement_results(
    returned_path: Path,
    manifest_path: Path,
    profile: CorrelationProfile,
    *,
    returned_sheet: str = REQUEST_SHEET,
    manifest_sheet: str = MANIFEST_SHEET,
) -> pd.DataFrame:
    """Validate returned IDs and merge only the CV value into the internal manifest."""
    returned_path, manifest_path = Path(returned_path), Path(manifest_path)
    returned_metadata = _read_metadata(returned_path)
    manifest_metadata = _read_metadata(manifest_path)
    for metadata, label in ((returned_metadata, "returned"), (manifest_metadata, "manifest")):
        if metadata.get("Schema_Version") != SCHEMA_VERSION:
            raise ValueError(f"The {label} workbook uses an unsupported handoff schema")
        if metadata.get("Profile") != profile.name:
            raise ValueError(f"The {label} workbook belongs to profile '{metadata.get('Profile')}'")

    returned = pd.read_excel(returned_path, sheet_name=returned_sheet)
    manifest = pd.read_excel(manifest_path, sheet_name=manifest_sheet)
    key = [REQUEST_ID, REPEAT_INDEX]
    required = {*key, profile.reference_column}
    missing = sorted(required - set(returned.columns))
    if missing:
        raise ValueError(f"Returned workbook is missing columns: {missing}")
    missing_manifest = sorted(set(key) - set(manifest.columns))
    if missing_manifest:
        raise ValueError(f"Manifest is missing columns: {missing_manifest}")
    if returned.duplicated(key).any():
        raise ValueError("Returned workbook contains duplicate measurement request keys")
    if manifest.duplicated(key).any():
        raise ValueError("Internal manifest contains duplicate measurement request keys")

    returned_keys = pd.MultiIndex.from_frame(returned[key])
    manifest_keys = pd.MultiIndex.from_frame(manifest[key])
    unknown = returned_keys.difference(manifest_keys)
    absent = manifest_keys.difference(returned_keys)
    if len(unknown) or len(absent):
        raise ValueError(f"Request coverage mismatch: {len(unknown)} unknown and {len(absent)} missing keys")

    values = returned[key + [profile.reference_column]].copy()
    values[profile.reference_column] = pd.to_numeric(values[profile.reference_column], errors="coerce")
    if values[profile.reference_column].isna().any():
        count = int(values[profile.reference_column].isna().sum())
        raise ValueError(f"Returned workbook contains {count} blank or non-numeric measurement values")
    if profile.reference_column in manifest.columns:
        manifest = manifest.drop(columns=profile.reference_column)
    return manifest.merge(values, on=key, how="left", validate="one_to_one")
