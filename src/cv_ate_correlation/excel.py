"""Consistent Excel workbook formatting for every CorreLaTE output."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, Mapping

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ACCENT_1_BLUE = "4472C4"
HEADER_FONT_COLOR = "FFFFFF"
DATA_INTEREST_FILL = "E2F0D9"
DATA_INTEREST_FONT_COLOR = "375623"
DEFAULT_MIN_WIDTH = 10
DEFAULT_MAX_WIDTH = 60


def format_worksheet(
    worksheet: Worksheet,
    *,
    min_width: int = DEFAULT_MIN_WIDTH,
    max_width: int = DEFAULT_MAX_WIDTH,
) -> None:
    """Apply Accent 1 headers, filters, frozen headers, and bounded autofit."""
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    if max_row < 1 or max_column < 1:
        return

    header_fill = PatternFill(fill_type="solid", fgColor=ACCENT_1_BLUE)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color=HEADER_FONT_COLOR, bold=True)
        cell.alignment = Alignment(vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(max_column)}{max_row}"
    worksheet.sheet_view.showGridLines = False

    longest_by_column = [0] * max_column
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_column):
        for column_offset, cell in enumerate(row):
            value = cell.value
            if value is None:
                continue
            lines = str(value).splitlines() or [""]
            longest_by_column[column_offset] = max(
                longest_by_column[column_offset],
                *(len(line) for line in lines),
            )
    for column_index, longest in enumerate(longest_by_column, start=1):
        width = min(max(longest + 2, min_width), max_width)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def highlight_data_columns(worksheet: Worksheet, column_names: Iterable[str]) -> None:
    """Highlight populated report cells whose headers identify primary outputs."""
    requested = set(column_names)
    if not requested or worksheet.max_row < 2:
        return
    fill = PatternFill(fill_type="solid", fgColor=DATA_INTEREST_FILL)
    font = Font(color=DATA_INTEREST_FONT_COLOR, bold=True)
    for column_index, header in enumerate(worksheet[1], start=1):
        if str(header.value or "") not in requested:
            continue
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if cell.value is None:
                continue
            cell.fill = fill
            cell.font = font


def format_workbook(
    path: str | Path,
    *,
    highlighted_columns: Mapping[str, Iterable[str]] | None = None,
) -> None:
    """Format every non-empty worksheet in an existing Excel workbook."""
    destination = Path(path)
    workbook = load_workbook(destination)
    for worksheet in workbook.worksheets:
        format_worksheet(worksheet)
        if highlighted_columns:
            highlight_data_columns(worksheet, highlighted_columns.get(worksheet.title, ()))
    workbook.save(destination)


def write_dataframe_workbook(
    path: str | Path,
    sheets: Mapping[str, pd.DataFrame],
) -> Path:
    """Write DataFrames and apply the standard CorreLaTE Excel presentation."""
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(destination, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, index=False, sheet_name=sheet_name)
        for worksheet in writer.book.worksheets:
            format_worksheet(worksheet)
    return destination
